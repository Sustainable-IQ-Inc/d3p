from fastapi import APIRouter, Depends
from fastapi.responses import FileResponse
import models
from utils import verify_token, add_event_history, supabase
from typing import Optional, Dict, Union
from uuid import uuid4
from gcs_upload import upload_blob, get_signed_url_from_url
import os
from multi_upload import process_multi_upload
import logging_start
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import tempfile
from google.cloud import storage

from post_processing import run_script_master


router = APIRouter()

def get_upload_status_id(status_name: str) -> Optional[int]:
    """Get upload status ID from enum_upload_statuses table"""
    try:
        data, _ = supabase.table('enum_upload_statuses')\
            .select('id')\
            .eq('name', status_name)\
            .limit(1)\
            .execute()
        if data and len(data) > 1 and data[1]:
            return data[1][0]['id']
    except Exception as e:
        logging_start.logger.error(f"Error getting upload status ID: {e}")
    return None

def get_user_email(user_id: str) -> Optional[str]:
    """Get user email from profiles table"""
    try:
        data, _ = supabase.table('profiles')\
            .select('email')\
            .eq('id', user_id)\
            .limit(1)\
            .execute()
        if data and len(data) > 1 and data[1]:
            return data[1][0].get('email')
    except Exception as e:
        logging_start.logger.error(f"Error getting user email: {e}")
    return None

def move_file_to_failed_folder(original_url: str, file_name: str) -> Optional[str]:
    """Move file from report_uploads/ to failed_uploads/ folder in GCS"""
    try:
        BUCKET_NAME = os.environ.get('BUCKET_NAME')
        if not BUCKET_NAME:
            return None
        
        # Parse the original URL to get blob name
        from urllib.parse import urlparse
        parsed = urlparse(original_url)
        # Extract blob name from path (remove leading /)
        original_blob_name = parsed.path.lstrip('/')
        
        # Extract just the filename part (after report_uploads/)
        if 'report_uploads/' in original_blob_name:
            filename_part = original_blob_name.split('report_uploads/')[1]
        else:
            filename_part = file_name
        
        # New blob name in failed_uploads folder
        new_blob_name = f'failed_uploads/{filename_part}'
        
        # Use storage client to copy and delete
        from gcs_upload import get_signing_client
        storage_client = get_signing_client()
        bucket = storage_client.bucket(BUCKET_NAME)
        
        source_blob = bucket.blob(original_blob_name)
        if not source_blob.exists():
            logging_start.logger.warning(f"Source blob {original_blob_name} does not exist")
            return None
        
        # Copy to new location
        new_blob = bucket.copy_blob(source_blob, bucket, new_blob_name)
        
        # Delete original
        source_blob.delete()
        
        # Generate signed URL for new location
        from datetime import timedelta
        url = new_blob.generate_signed_url(
            version="v4",
            expiration=timedelta(hours=1),
            method="GET",
        )
        
        logging_start.logger.info(f"File moved from {original_blob_name} to {new_blob_name}")
        return url
        
    except Exception as e:
        logging_start.logger.error(f"Error moving file to failed folder: {e}")
        # If move fails, return original URL so we still have a reference
        return original_url

def create_failed_upload_record(
    file_url: str,
    file_name: str,
    error_message: str,
    user_id: str,
    company_id: str,
    baseline_design: Optional[str] = None,
    report_type_id: Optional[int] = None
) -> Optional[int]:
    """Create an upload record with failed status"""
    try:
        failed_status_id = get_upload_status_id('failed')
        if not failed_status_id:
            logging_start.logger.error("Failed status not found in enum_upload_statuses")
            return None
        
        upload_data = {
            'file_url': file_url,
            'file_name': file_name,
            'processing_error': error_message,
            'upload_status_id': failed_status_id,
            'user_id': user_id,
            'company_id': company_id,
            'notified_admin': False,
            'notified_user_complete': False
        }
        
        if baseline_design:
            upload_data['baseline_status' if baseline_design == 'baseline' else 'design_status'] = 'failed'
            upload_data['design_baseline_type'] = baseline_design
        
        if report_type_id:
            upload_data['report_type_id'] = report_type_id
        
        data, _ = supabase.table('uploads')\
            .insert(upload_data)\
            .execute()
        
        if data and len(data) > 1 and data[1]:
            upload_id = data[1][0]['id']
            logging_start.logger.info(f"Created failed upload record {upload_id}")
            return upload_id
            
    except Exception as e:
        logging_start.logger.error(f"Error creating failed upload record: {e}")
    return None

def update_eeu_record(eeu_id, upload_id):
            eeu_data_dict  = dict()
            eeu_data_dict['id'] = eeu_id
            eeu_data_dict['upload_id'] = upload_id

            try:
                data, count = supabase.table('eeu_data')\
                    .update(eeu_data_dict)\
                    .eq('id', eeu_id)\
                    .execute()
            except Exception as e:
                print(e)
                return "error eeu table"
            return "success"

def upload_report(url, baseline_design, report_type=None, conditioned_area=None, file_extension=None, file_name=None, user_id=None, company_id=None):
    print(f"DEBUG: upload_report called with url={url}, baseline_design={baseline_design}, report_type={report_type}, file_extension={file_extension}")
    
    # Check if this is a multi-project Excel file BEFORE running script master
    if file_extension and file_extension.lower() in ['.xlsx', '.xls']:
        try:
            from parse_reports.parse_multi_project_xlsx import is_multi_project_excel
            if is_multi_project_excel(url):
                logging_start.logger.info(f"Multi-project Excel file detected, skipping report type detection")
                # Return special indicator for multi-project files
                return {
                    'status': 'success',
                    'report_type': 9,
                    'is_multi_project': True,
                    'file_url': url,
                    'message': "Multi-project Excel file detected - use multi-project service"
                }
        except Exception as e:
            logging_start.logger.info(f"Multi-project detection failed: {str(e)}, proceeding with normal parsing")
    
    args = {
        'url': url,
        'conditioned_area': conditioned_area,
        'baseline_design': baseline_design
    }
    if report_type is not None:
        args['report_type'] = report_type

    print(f"DEBUG: Calling run_script_master with args: {args}")
    results = run_script_master(**args)
    print(f"DEBUG: run_script_master returned: {type(results)} - {results}")
    
    # Check if this is a multi-project Excel file (report type 9)
    if (isinstance(results, dict) and 
        results.get('status') == 'success' and 
        results.get('report_type') == 9):
        # Return special indicator for multi-project files
        return {
            'status': 'success',
            'report_type': 9,
            'projects': results.get('projects', []),
            'validation_errors': results.get('validation_errors', []),
            'message': f"Multi-project Excel file detected with {len(results.get('projects', []))} projects"
        }
    
    # Handle both dictionary and list return formats from run_script_master
    if isinstance(results, list):
        print(f"DEBUG: Results is a list: {results}")
        # Legacy format: ["status", errors, warnings] or ["ERROR", errors, warnings]
        status = results[0]
        errors = results[1] if len(results) > 1 else []
        warnings = results[2] if len(results) > 2 else []
        
        if status in ["pending", "ERROR"]:
            print(f"ERROR: Processing failed with status: {status}, errors: {errors}")
            errors_flat = [item for sublist in errors for item in sublist] if errors and len(errors) > 0 and isinstance(errors[0], list) else errors
            warnings_flat = [item for sublist in warnings for item in sublist] if warnings and len(warnings) > 0 and isinstance(warnings[0], list) else warnings
            error_msg = '\n'.join(errors_flat) if errors_flat else f'File processing failed with status: {status}'
            # If user_id and company_id provided, create failed upload record
            if user_id and company_id:
                logging_start.logger.info(f"Creating failed upload record for file {file_name}, user_id: {user_id}, company_id: {company_id}")
                failed_url = move_file_to_failed_folder(url, file_name)
                if failed_url:
                    upload_id = create_failed_upload_record(
                        failed_url, file_name, error_msg, user_id, company_id,
                        baseline_design, report_type
                    )
                    if upload_id:
                        logging_start.logger.info(f"Successfully created failed upload record {upload_id}")
                    else:
                        logging_start.logger.error(f"Failed to create upload record for {file_name}")
                else:
                    logging_start.logger.error(f"Failed to move file {file_name} to failed_uploads folder")
            else:
                logging_start.logger.warning(f"Not creating failed upload record - missing user_id or company_id. user_id: {user_id}, company_id: {company_id}")
            return {
                'status': 'error',
                'errors': error_msg,
                'warnings': '\n'.join(warnings_flat) if warnings_flat else '',
                'message': f'File processing failed with status: {status}'
            }
    else:
        print(f"DEBUG: Results is a dict: {results}")
        # Dictionary format: {"status": "success", "df": df, "errors": [], "warnings": [], "report_type": int}
        if results.get('status') != 'success':
            print(f"ERROR: Results status is not success: {results.get('status')}")
            error_msg = str(results.get('errors', []))
            # If user_id and company_id provided, create failed upload record
            if user_id and company_id:
                failed_url = move_file_to_failed_folder(url, file_name)
                if failed_url:
                    create_failed_upload_record(
                        failed_url, file_name, error_msg, user_id, company_id,
                        baseline_design, report_type
                    )
            return {
                'status': 'error',
                'message': 'File processing failed',
                'errors': error_msg,
                'warnings': str(results.get('warnings', []))
            }
        
        warnings = results['warnings']
        errors = results['errors']
    
    # Ensure we have a dataframe to work with
    if isinstance(results, list) or 'df' not in results:
        print(f"ERROR: No valid dataframe found in results: {results}")
        error_msg = 'No valid data found in file'
        # If user_id and company_id provided, create failed upload record
        if user_id and company_id:
            failed_url = move_file_to_failed_folder(url, file_name)
            if failed_url:
                create_failed_upload_record(
                    failed_url, file_name, error_msg, user_id, company_id,
                    baseline_design, report_type
                )
        return {
            'status': 'error',
            'message': error_msg,
            'errors': str(errors) if 'errors' in locals() else '',
            'warnings': str(warnings) if 'warnings' in locals() else ''
        }
    
    warnings_flat = [item for sublist in warnings for item in sublist] if warnings and len(warnings) > 0 and isinstance(warnings[0], list) else warnings
    warnings_str = '\n'.join(warnings_flat) if warnings_flat else ''

    errors_flat = [item for sublist in errors for item in sublist] if errors and len(errors) > 0 and isinstance(errors[0], list) else errors
    errors_str = '\n'.join(errors_flat) if errors_flat else ''
    
    print(f"DEBUG: Converting DataFrame to dict. DataFrame shape: {results['df'].shape}")
    print(f"DEBUG: DataFrame columns: {list(results['df'].columns)}")
    print(f"DEBUG: DataFrame dtypes: {results['df'].dtypes}")
    
    # Clean the DataFrame before converting to dict
    df_clean = results['df'].copy()
    
    # Remove invalid column names (like numeric column names)
    invalid_columns = [col for col in df_clean.columns if isinstance(col, (int, float)) or str(col).replace('.', '').isdigit()]
    if invalid_columns:
        print(f"DEBUG: Removing invalid columns: {invalid_columns}")
        df_clean = df_clean.drop(columns=invalid_columns)
    
    # Replace NaN values with None (which becomes null in JSON)
    df_clean = df_clean.where(pd.notnull(df_clean), None)
    
    # Convert to dict
    eeu_data_to_insert = df_clean.to_dict(orient='records')
    print(f"DEBUG: Converted to dict, length: {len(eeu_data_to_insert)}")
    print(f"DEBUG: First record keys: {list(eeu_data_to_insert[0].keys()) if eeu_data_to_insert else 'No records'}")
    
    # Only process the first record (the actual data, not the header row)
    if len(eeu_data_to_insert) > 1:
        eeu_data_to_insert = [eeu_data_to_insert[1]]  # Take the second record (index 1) which has the actual data
    
    eeu_data_to_insert[0]['baseline_design'] = baseline_design
    eeu_data_to_insert[0]['upload_warnings'] = warnings_str
    eeu_data_to_insert[0]['upload_errors'] = errors_str
    eeu_data_to_insert[0]['file_type'] = file_extension
    eeu_data_to_insert[0]['file_name'] = file_name
    eeu_data_to_insert[0]['file_url'] = url
    
    print(f"DEBUG: Final data to insert: {eeu_data_to_insert}")
    
    try:
        data, count = supabase.table('eeu_data')\
            .insert(eeu_data_to_insert)\
            .execute()
        print(f"DEBUG: Successfully inserted data, count: {count}")
    except Exception as e:
        print(f"ERROR: Database insertion failed: {e}")
        print(f"ERROR: Data that failed to insert: {eeu_data_to_insert}")
        error_msg = f"Database insertion failed: {str(e)}"
        # If user_id and company_id provided, create failed upload record
        if user_id and company_id:
            failed_url = move_file_to_failed_folder(url, file_name)
            if failed_url:
                create_failed_upload_record(
                    failed_url, file_name, error_msg, user_id, company_id,
                    baseline_design, results.get('report_type')
                )
        return {
            'status': 'error',
            'errors': error_msg,
            'warnings': warnings_str,
            'message': 'File processing failed during database insertion'
        }
    avg_energy = float(eeu_data_to_insert[0]['total_energy'])*1000 / float(eeu_data_to_insert[0]['use_type_total_area'])

    return {'status': "success",
            'eeu_id': data[1][0]['id'],
            'conditioned_area': eeu_data_to_insert[0]['use_type_total_area'],
            'climate_zone': eeu_data_to_insert[0]['climate_zone'],
            'total_energy': eeu_data_to_insert[0]['total_energy'],
            'avg_energy': avg_energy,
            'url': url,
            'errors': errors_str,
            'warnings': warnings_str,
            'file_name': file_name,
            'report_type': results.get('report_type')
            }
    
@router.post("/uploadfile/")
async def create_upload_file(item: models.ReportUpload = Depends(), authorized: Dict[str, Union[bool, Optional[str]]] = Depends(verify_token)):
    print(f"DEBUG: Upload request received. Authorized: {authorized['is_authorized']}")
    if authorized['is_authorized']:
        print(f"DEBUG: File info - filename: {item.file.filename}, baseline_design: {item.baseline_design}")
        
        BUCKET_NAME = os.environ.get('BUCKET_NAME')

        if BUCKET_NAME is None:
            print("ERROR: BUCKET_NAME environment variable is not set")
            raise ValueError("BUCKET_NAME environment variable is not set")
        #create a unique uuid for the filename
        uuid = uuid4()
        filename_new = 'report_uploads/' + str(uuid) + item.file.filename
        file_name = item.file.filename
        file_extension = os.path.splitext(filename_new)[1]
        print(f"DEBUG: Uploading file to GCS - filename: {filename_new}, extension: {file_extension}")
        url = upload_blob(BUCKET_NAME, filename_new, file_obj = item.file.file,)
        print(f"DEBUG: File uploaded to GCS, URL: {url}")

        

        #if item.report_type == 8:
        if hasattr(item, 'report_type') and item.report_type == 8:
            report_type = item.report_type
            user_id = authorized.get('user_id')
            company_id = authorized.get('company_id')
            baseline_output = upload_report(url,"baseline",report_type,file_extension=file_extension,file_name = file_name, user_id=user_id, company_id=company_id)
            design_output = upload_report(url,"design",report_type,file_extension=file_extension, file_name = file_name, user_id=user_id, company_id=company_id)
            
            # Handle partial failures - if one side fails, still allow form completion
            response = {'report_type': report_type, 'baseline': baseline_output, 'design': design_output}
            if baseline_output.get('status') == 'error' or design_output.get('status') == 'error':
                response['allow_form_completion'] = True
                response['message'] = 'One or both sides could not be processed automatically. You can still complete the form below.'
            
            return response

        else:
            user_id = authorized.get('user_id')
            company_id = authorized.get('company_id')
            args = {
                'url': url,
                'baseline_design': item.baseline_design,
                'conditioned_area': item.conditioned_area,
                'file_extension': file_extension,
                'file_name': file_name,
                'user_id': user_id,
                'company_id': company_id
            }
            if hasattr(item, 'report_type') and item.report_type is not None:
                args['report_type'] = item.report_type

            report_to_upload = upload_report(**args)
            print(f"DEBUG: upload_report returned: {report_to_upload}")
            
            # Check if the parsed report type is PRM (type 8), and if so, process both baseline and design
            if (isinstance(report_to_upload, dict) and 
                'status' in report_to_upload and 
                report_to_upload['status'] == 'success' and 
                report_to_upload.get('report_type') == 8):  # PRM report
                
                # Process both baseline and design for PRM reports
                baseline_output = upload_report(url, "baseline", report_type=8, file_extension=file_extension, file_name=file_name, user_id=user_id, company_id=company_id)
                design_output = upload_report(url, "design", report_type=8, file_extension=file_extension, file_name=file_name, user_id=user_id, company_id=company_id)
                
                # Handle partial failures - if one side fails, still allow form completion
                response = {
                    'report_type': 8,
                    'baseline': baseline_output,
                    'design': design_output
                }
                if baseline_output.get('status') == 'error' or design_output.get('status') == 'error':
                    response['allow_form_completion'] = True
                    response['message'] = 'One or both sides could not be processed automatically. You can still complete the form below.'
                
                return response
            
            # Handle failed uploads - return error but allow form completion
            if isinstance(report_to_upload, dict) and report_to_upload.get('status') == 'error':
                # Return error response but indicate form can still be completed
                return {
                    'status': 'error',
                    'message': 'File could not be processed automatically. You can still complete the form below. We\'ll notify you when processing is complete.',
                    'errors': report_to_upload.get('errors', ''),
                    'warnings': report_to_upload.get('warnings', ''),
                    'file_url': url,
                    'file_name': file_name,
                    'allow_form_completion': True
                }
            
            # Check if the parsed report type is Multi-Project Excel (type 9)
            if (isinstance(report_to_upload, dict) and 
                'status' in report_to_upload and 
                report_to_upload['status'] == 'success' and 
                report_to_upload.get('report_type') == 9):  # Multi-Project Excel
                
                # Use multi-project service to process the file
                try:
                    from multi_project_service import create_multi_project_service
                    service = create_multi_project_service()
                    result = service.process_multi_project_excel(url, item.company_id)
                    
                    return {
                        'status': result['status'],
                        'report_type': 9,
                        'total_projects': result['total_projects'],
                        'successful_projects': result['successful_projects'],
                        'failed_projects': result['failed_projects'],
                        'validation_errors': result['validation_errors'],
                        'created_project_ids': result.get('created_project_ids', []),
                        'created_projects': result.get('created_projects', []),
                        'message': f"Processed {result['successful_projects']} of {result['total_projects']} projects successfully"
                    }
                except Exception as e:
                    logging_start.logger.error(f"Error processing multi-project Excel: {str(e)}")
                    return {
                        'status': 'error',
                        'message': f"Multi-project processing failed: {str(e)}",
                        'report_type': 9
                    }
            
            print(f"DEBUG: Returning final result: {report_to_upload}")
            return report_to_upload

    else:
        print("ERROR: User not authorized")
        return "not authorized"
@router.post("/submit_multi_upload/")
async def submit_multiupload(item: models.MultiUpload, authorized: Dict[str, Union[bool, Optional[str]]] = Depends(verify_token)):
    if authorized['is_authorized']:
        company_id = authorized['company_id']
        try:
            multi_upload_response = process_multi_upload(company_id,item.design_files,item.baseline_files)
        except Exception as e:
            print(e)
            return "error multi upload"
        return multi_upload_response
        
    else:
        return "not authorized"

@router.post("/submit_multi_project_excel/")
async def submit_multi_project_excel(item: models.MultiProjectExcelUpload, authorized: Dict[str, Union[bool, Optional[str]]] = Depends(verify_token)):
    """
    Process multi-project Excel file and create multiple projects
    """
    if not authorized['is_authorized']:
        return {"status": "error", "message": "not authorized"}
    
    try:
        from multi_project_service import create_multi_project_service
        
        service = create_multi_project_service()
        result = service.process_multi_project_excel(item.file_url, item.company_id)
        
        return models.MultiProjectResult(
            status=result['status'],
            total_projects=result['total_projects'],
            successful_projects=result['successful_projects'],
            failed_projects=result['failed_projects'],
            validation_errors=result['validation_errors'],
            created_project_ids=result.get('created_project_ids', []),
            created_projects=result.get('created_projects', [])
        )
        
    except Exception as e:
        logging_start.logger.error(f"Error in multi-project Excel upload: {str(e)}")
        return {
            "status": "error", 
            "message": f"Processing failed: {str(e)}",
            "total_projects": 0,
            "successful_projects": 0,
            "failed_projects": 0,
            "validation_errors": [str(e)],
            "created_project_ids": []
        }

@router.get("/download-multi-project-template/")
async def download_multi_project_template():
    """
    Download the multi-project Excel template with baseline/design energy field columns
    """
    template_path = os.path.join(os.path.dirname(__file__), "dependencies", "d3p-multi-project-template.xlsx")
    
    if not os.path.exists(template_path):
        return {"error": "Template file not found"}

    try:
        # Load template workbook
        wb = load_workbook(template_path)
        ws = wb.active

        # Locate header row by scanning first few rows
        header_row_idx = None
        target_headers = {"project_name", "conditioned_area_sf", "project_use_type"}
        for row_idx in range(1, 6):
            values = [cell.value if cell.value is not None else "" for cell in ws[row_idx]]
            value_set = {str(v).strip() for v in values if str(v).strip() != ""}
            if target_headers.issubset(value_set):
                header_row_idx = row_idx
                break
        if header_row_idx is None:
            header_row_idx = 1

        # Define the EEU energy fields that need baseline/design suffixes
        eeu_energy_fields = [
            'Heating_Electricity', 'Heating_NaturalGas', 'Heating_DistrictHeating', 'Heating_Other',
            'Cooling_Electricity', 'Cooling_DistrictHeating', 'Cooling_Other',
            'DHW_Electricity', 'DHW_NaturalGas', 'DHW_DistrictHeating', 'DHW_Other',
            'Interior Lighting_Electricity', 'Exterior Lighting_Electricity', 'Plug Loads_Electricity',
            'Process Refrigeration_Electricity', 'Fans_Electricity', 'Pumps_Electricity', 'Pumps_NaturalGas',
            'Heat Rejection_Electricity', 'Humidification_Electricity', 'HeatRecovery_Electricity', 'HeatRecovery_Other',
            'ExteriorUsage_Electricity', 'ExteriorUsage_NaturalGas', 'OtherEndUse_Electricity', 'OtherEndUse_NaturalGas', 'OtherEndUse_Other',
            'SolarDHW_On-SiteRenewables', 'SolarPV_On-SiteRenewables', 'Wind_On-SiteRenewables', 'Other_On-SiteRenewables'
        ]

        # Create new headers with baseline/design suffixes for energy fields
        new_headers = []
        
        # First, add all the non-energy fields (shared fields)
        for col_idx, cell in enumerate(ws[header_row_idx], start=1):
            if cell.value:
                header_name = str(cell.value).strip()
                if header_name not in eeu_energy_fields:
                    new_headers.append(header_name)
        
        # Then add all baseline energy fields
        for energy_field in eeu_energy_fields:
            new_headers.append(f"{energy_field}_baseline")
        
        # Finally add all design energy fields
        for energy_field in eeu_energy_fields:
            new_headers.append(f"{energy_field}_design")

        # Clear the existing header row and add new headers
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=header_row_idx, column=col_idx).value = None
        
        for col_idx, header in enumerate(new_headers, start=1):
            ws.cell(row=header_row_idx, column=col_idx).value = header
        
        # Apply formatting to the header row
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Style the header row
        for col_idx in range(1, len(new_headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
            # Color code the columns
            header_name = new_headers[col_idx - 1]
            if header_name.endswith('_baseline'):
                cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')  # Light blue
            elif header_name.endswith('_design'):
                cell.fill = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')  # Light green
            else:
                cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')  # Light gray for shared fields

        # Update instruction text to reflect new structure
        ws.cell(row=2, column=1).value = "Include energy data in MBTU. Enter baseline values in blue columns, design values in green columns."
        ws.cell(row=3, column=1).value = "Enter 0.0 for energy types that do not apply to your project."
        
        # Apply borders only to cells with data in instruction rows
        for row_idx in [1, 2, 3]:
            for col_idx in range(1, len(new_headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Only apply borders to cells that have data
                if cell.value is not None and str(cell.value).strip() != '':
                    cell.border = thin_border
                    if col_idx == 1:  # Only the first column has text
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Update sample data rows to have clean baseline and design values
        shared_fields_count = len([h for h in new_headers if not h.endswith('_baseline') and not h.endswith('_design')])
        
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            # Clear all existing data beyond shared fields
            for col_idx in range(shared_fields_count + 1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).value = None
            
            # Add clean sample data for baseline energy fields
            baseline_start_col = shared_fields_count + 1
            for i, energy_field in enumerate(eeu_energy_fields):
                col_idx = baseline_start_col + i
                # Use 0.0 for most fields, with a few non-zero examples
                if i < 3:  # First 3 fields get sample values
                    ws.cell(row=row_idx, column=col_idx).value = 100.0 + (i * 50)
                else:
                    ws.cell(row=row_idx, column=col_idx).value = 0.0
            
            # Add clean sample data for design energy fields
            design_start_col = baseline_start_col + len(eeu_energy_fields)
            for i, energy_field in enumerate(eeu_energy_fields):
                col_idx = design_start_col + i
                # Use 0.0 for most fields, with a few non-zero examples
                if i < 3:  # First 3 fields get sample values (typically 20% better than baseline)
                    ws.cell(row=row_idx, column=col_idx).value = 80.0 + (i * 40)
                else:
                    ws.cell(row=row_idx, column=col_idx).value = 0.0
            
            # Apply formatting to data rows
            for col_idx in range(1, len(new_headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
                # Apply background color to data cells
                header_name = new_headers[col_idx - 1]
                if header_name.endswith('_baseline'):
                    cell.fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')  # Very light blue
                elif header_name.endswith('_design'):
                    cell.fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')  # Very light green

        # Map header names to column letters for validation
        headers = {}
        for col_idx, header in enumerate(new_headers, start=1):
            headers[header] = get_column_letter(col_idx)

        # Prepare allowed values via database and constants
        def fetch_enum_names(table_name: str) -> list:
            try:
                data, _ = supabase.table(table_name).select('name').order('order').execute()
                if data and len(data) > 1:
                    return [row['name'] for row in data[1] if row.get('name')]
            except Exception as e:
                logging_start.logger.warning(f"Failed to fetch {table_name}: {str(e)}")
            return []

        allowed_values_map = {
            'project_use_type': fetch_enum_names('enum_project_use_types'),
            'project_construction_category': fetch_enum_names('enum_project_construction_categories'),
            'project_phase': fetch_enum_names('enum_project_phases'),
            'energy_code': fetch_enum_names('enum_energy_codes'),
            'report_type': fetch_enum_names('enum_report_types'),
            'climate_zone': fetch_enum_names('enum_climate_zones'),
            'area_units': ['sf', 'sm'],
            'energy_units': ['mbtu', 'gj'],
        }

        # Create or replace "Valid Values" sheet
        if 'Valid Values' in wb.sheetnames:
            del wb['Valid Values']
        valid_ws = wb.create_sheet('Valid Values')

        # Write allowed values, one field per column
        field_names = list(allowed_values_map.keys())
        for col_idx, field in enumerate(field_names, start=1):
            col_letter = get_column_letter(col_idx)
            valid_ws[f"{col_letter}1"] = field
            for row_offset, value in enumerate(allowed_values_map[field], start=2):
                valid_ws[f"{col_letter}{row_offset}"] = value

        # Add data validations to first 100 rows after header
        max_rows = 100
        for field, values in allowed_values_map.items():
            if field in headers and values:
                col_letter = headers[field]
                end_row = 1 + len(values)
                formula = f"='Valid Values'!${get_column_letter(field_names.index(field)+1)}$2:${get_column_letter(field_names.index(field)+1)}${end_row}"
                dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=True)
                dv.error = "Select a value from the list."
                dv.errorTitle = "Invalid value"
                ws.add_data_validation(dv)
                for r in range(header_row_idx + 1, header_row_idx + 1 + max_rows):
                    dv.add(f"{col_letter}{r}")

        # Save to a temp file and serve
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp_path = tmp.name
        tmp.close()
        wb.save(tmp_path)

        return FileResponse(
            path=tmp_path,
            filename="d3p-multi-project-template.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logging_start.logger.error(f"Error generating template with validations: {str(e)}")
        # Fall back to static template if dynamic generation fails
        return FileResponse(
            path=template_path,
            filename="d3p-multi-project-template.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

@router.post("/submit_project/")
async def submit_project(item: models.SubmitProject, authorized: Dict[str, Union[bool, Optional[str]]] = Depends(verify_token)):
    

    if authorized['is_authorized']:
        user_id = authorized.get('user_id')
        company_id = authorized.get('company_id')
        
        # Convert the item to a dictionary
        item_data = item.model_dump()
        
        item_data.setdefault('design_eeu_id', None)
        item_data.setdefault('energy_code_id', None)
        
        # Always set company_id (user_id is bigint in uploads table, but we have UUID from auth)
        # The user_id field in uploads references the old users table, so we'll leave it NULL
        # and rely on company_id for filtering
        item_data['company_id'] = company_id
        # Don't set user_id - it's a bigint foreign key to users table, not UUID from auth

        extract_dict = {}

        if 'baseline_eeu_id' in item_data:
            extract_dict['baseline_eeu_id'] = item_data.pop('baseline_eeu_id')

        if 'design_eeu_id' in item_data:
            extract_dict['design_eeu_id'] = item_data.pop('design_eeu_id')
        
        # Check for failed uploads that need admin notification and potentially update
        failed_upload_id = None
        try:
            failed_status_id = get_upload_status_id('failed')
            if failed_status_id:
                # Look for a recent failed upload that matches this submission
                # (within last 24 hours, same user/company, no project_id yet)
                from datetime import datetime, timedelta
                recent_time = (datetime.utcnow() - timedelta(hours=24)).isoformat()
                
                # Note: user_id in uploads is bigint, but we have UUID from auth, so filter by company_id only
                failed_uploads_query = supabase.table('uploads')\
                    .select('id, file_name, processing_error, baseline_status, design_status, notified_admin, project_id')\
                    .eq('upload_status_id', failed_status_id)\
                    .eq('company_id', company_id)\
                    .is_('project_id', 'null')\
                    .gte('created_at', recent_time)\
                    .order('created_at', desc=True)\
                    .limit(1)
                
                failed_uploads_data, _ = failed_uploads_query.execute()
                
                # If we found a matching failed upload, update it instead of creating new
                if failed_uploads_data and len(failed_uploads_data) > 1 and failed_uploads_data[1]:
                    failed_upload = failed_uploads_data[1][0]
                    failed_upload_id = failed_upload['id']
                    
                    # Update the failed upload record with the new data
                    # Set status to pending since user has now provided project details
                    # Preserve file_name and file_url from the original failed upload
                    pending_status_id = get_upload_status_id('pending')
                    update_data = {
                        **item_data,
                        'upload_status_id': pending_status_id if pending_status_id else failed_status_id,
                        'processing_error': None,  # Clear error since user is trying again
                        # Preserve file info from original failed upload
                        'file_name': failed_upload.get('file_name') or item_data.get('file_name'),
                        'file_url': failed_upload.get('file_url') or item_data.get('file_url')
                    }
                    
                    supabase.table('uploads')\
                        .update(update_data)\
                        .eq('id', failed_upload_id)\
                        .execute()
                    
                    logging_start.logger.info(f"Updated existing failed upload {failed_upload_id} with project details")
                
                # Send notifications for any failed uploads that haven't been notified yet
                # Note: user_id in uploads is bigint, but we have UUID from auth, so filter by company_id only
                all_failed_query = supabase.table('uploads')\
                    .select('id, file_name, processing_error, baseline_status, design_status, notified_admin')\
                    .eq('upload_status_id', failed_status_id)\
                    .eq('company_id', company_id)\
                    .eq('notified_admin', False)
                
                all_failed_data, _ = all_failed_query.execute()
                
                if all_failed_data and len(all_failed_data) > 1:
                    user_email = get_user_email(user_id)
                    # Get company name
                    company_name = None
                    try:
                        company_data, _ = supabase.table('companies')\
                            .select('company_name')\
                            .eq('id', company_id)\
                            .limit(1)\
                            .execute()
                        if company_data and len(company_data) > 1 and company_data[1]:
                            company_name = company_data[1][0].get('company_name')
                    except:
                        pass
                    
                    for failed_upload in all_failed_data[1]:
                        # Skip if this is the one we just updated
                        if failed_upload['id'] == failed_upload_id:
                            continue
                            
                        baseline_design = None
                        if failed_upload.get('baseline_status') == 'failed':
                            baseline_design = 'baseline'
                        elif failed_upload.get('design_status') == 'failed':
                            baseline_design = 'design'
                        
                        from email_service import send_failed_upload_notification_to_admin
                        send_failed_upload_notification_to_admin(
                            failed_upload['id'],
                            user_email or 'unknown',
                            failed_upload.get('file_name', 'unknown'),
                            failed_upload.get('processing_error', 'Unknown error'),
                            baseline_design,
                            company_name
                        )
                        
                        # Mark as notified
                        supabase.table('uploads')\
                            .update({'notified_admin': True})\
                            .eq('id', failed_upload['id'])\
                            .execute()
        except Exception as e:
            logging_start.logger.error(f"Error checking failed uploads: {e}")

        try:
            # If we updated an existing failed upload, use that ID; otherwise create new
            if failed_upload_id:
                data = (None, [{'id': failed_upload_id}])
            else:
                # For new uploads, set status to pending if no eeu_ids provided (file processing pending)
                # or completed if eeu_ids are provided (file already processed)
                if not extract_dict.get('baseline_eeu_id') and not extract_dict.get('design_eeu_id'):
                    pending_status_id = get_upload_status_id('pending')
                    if pending_status_id:
                        item_data['upload_status_id'] = pending_status_id
                else:
                    completed_status_id = get_upload_status_id('completed')
                    if completed_status_id:
                        item_data['upload_status_id'] = completed_status_id
                
                logging_start.logger.info(f"Inserting upload record with data: {item_data}")
                data, count = supabase.table('uploads')\
                    .insert(item_data)\
                    .execute()
                logging_start.logger.info(f"Successfully inserted upload record: {data[1][0]['id'] if data and len(data) > 1 and data[1] else 'No data returned'}")
        except Exception as e:
            error_msg = f"Error inserting upload record: {str(e)}"
            logging_start.logger.error(error_msg, exc_info=True)
            print(error_msg)
            return {"error": error_msg, "status": "failed"}
        
        def update_eeu_if_exists(key, extract_dict, upload_id):
            if extract_dict.get(key) is not None:
                eeu_data_dict = {
                    'id': extract_dict[key],
                    'upload_id': upload_id
                }
                update_eeu_record(eeu_data_dict['id'], eeu_data_dict['upload_id'])

        # Usage - only update if eeu_ids exist (may be null for failed uploads)
        if not data or len(data) <= 1 or not data[1]:
            error_msg = "No upload record returned after insert/update"
            logging_start.logger.error(error_msg)
            return {"error": error_msg, "status": "failed"}
            
        upload_id = data[1][0]['id']
        update_eeu_if_exists('baseline_eeu_id', extract_dict, upload_id)
        update_eeu_if_exists('design_eeu_id', extract_dict, upload_id)
        
        logging_start.logger.info(f"Successfully created/updated upload {upload_id} for project {item_data.get('project_id')}")

        return "success"

    else:
        return "not authorized"